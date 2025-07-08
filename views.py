from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render, redirect
from .forms import *
from .models import *
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import io, base64, csv
import seaborn as sns
import pandas as pd
from io import BytesIO
import matplotlib

matplotlib.use('Agg')


def index(request):
    return render(request, 'index.html', {})


def admin_edit_coupon(request, id):
    x = Admin_coupon.objects.get(id=id)
    if request.method == "POST":
        form = Admin_couponForm(request.POST, request.FILES, instance=x)
        print(form.errors)
        if form.is_valid():
            form.save()
            return redirect('/view_')
        return render(request, 'customer_update_.html',
                      {"x": x, 'msg': 'Your  Is Not Updated'})
    return render(request, 'customer_update_.html', {"x": x})


def admin_add_coupon(request):
    if request.method == "POST":
        form = Admin_couponForm(request.POST)
        print(form.errors)
        if form.is_valid():
            form.save()
            return render(request, "admin_add_coupon.html", {'msg': 'Added Coupon Successfully'})
        return render(request, "admin_add_coupon.html", {})
    return render(request, "admin_add_coupon.html", {})


def admin_view_coupon(request):
    coupon = Admin_coupon.objects.all()
    return render(request, 'admin_view_coupon.html', {"coupon": coupon})


def del_coupon(request, id):
    cus = Admin_coupon.objects.get(id=id)
    cus.delete()
    return redirect('/admin_view_coupon')


def admin_delete_not(request, id):
    cus = Notifications.objects.get(id=id)
    cus.delete()
    return redirect('/admin_view_notification')


def del_customer(request, id):
    cus = Registration.objects.get(id=id)
    cus.status = 'On Hold'
    cus.save()
    return redirect('/registration')





def customer_view_details(request, id):
    cudet = _add_.objects.get(id=id)
    return render(request, 'customer_view_details.html', {"x": cudet})






def view_positive_speech(request):
    obj = SpeechModel.objects.filter(status="positive speech")
    return render(request, 'view_positive_speech.html', {"obj": obj})


def view_negative_speech(request):
    obj = SpeechModel.objects.filter(status="negative speech")
    return render(request, 'view_negative_speech.html', {"obj": obj})


def view_hate_speech(request):
    obj = SpeechModel.objects.filter(status="hate speech")
    return render(request, 'view_hate_speech.html', {"obj": obj})


def add_speech(request):
    email = request.session["email"]
    obj1 = Registration.objects.get(email=email)
    if request.method == "POST":
        form = SpeechForm(request.POST)
        print(form.errors)
        if form.is_valid():
            user = form.cleaned_data["user"]
            names = form.cleaned_data["names"]
            dob = form.cleaned_data["dob"]
            age = form.cleaned_data["age"]
            gender = form.cleaned_data["gender"]
            ethnicity = form.cleaned_data["ethnicity"]
            religion = form.cleaned_data["religion"]
            color = form.cleaned_data["color"]
            nationality = form.cleaned_data["nationality"]
            city = form.cleaned_data["city"]
            speecher = form.cleaned_data["speecher"]
            location = form.cleaned_data["location"]
            speech = form.cleaned_data["speech"]

            # Calculate sentiment polarity and label
            res = TextBlob(speech)
            polarity = res.sentiment.polarity
            sentiment_label = get_sentiment(speech)

            print(f"Polarity: {polarity}, Sentiment: {sentiment_label}")

            obj = Speech()
            obj.user = user
            obj.names = names
            obj.gender = gender
            obj.ethnicity = ethnicity
            obj.religion = religion
            obj.color = color
            obj.nationality = nationality
            obj.city = city
            obj.location = location
            obj.speecher = speecher
            obj.age = age
            obj.score = polarity  # Save the polarity as the score
            obj.dob = dob
            obj.speech = speech
            obj.status = sentiment_label  # Save the sentiment label as the status
            obj.save()
            csv_file = "data.csv"
            with open(csv_file, mode="a", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                file.seek(0, 2)
                if file.tell() == 0:
                    writer.writerow(['User', 'Name', 'DOB', 'Age', 'Gender', 'Ethnicity', 'Religion',
                                     'Color', 'Nationality', 'City', 'Speecher', 'Location',
                                     'Speech'])
                writer.writerow([user.id, names, dob, age, gender, ethnicity, religion,
                                 color, nationality, city, speecher, location,
                                 speech])
            return render(request, "add_speech.html",
                          {'msg': 'Your Speech Added Successfully', "email": email, "user": obj1})
        return render(request, "add_speech.html", {'msg': 'Your Speech Not Added', "email": email, "user": obj1})
    return render(request, "add_speech.html", {"email": email, "user": obj1})


def upload_dataset(request):
    if request.method == "POST" and request.FILES["csv_file"]:

        # Get the uploaded CSV file
        csv_file = request.FILES["csv_file"]

        # Prepare to store analyzed data
        analyzed_data = []

        # First, delete all existing records in the table to start fresh
        SpeechModel.objects.all().delete()

        # Open and read the CSV file
        decoded_file = csv_file.read().decode("utf-8").splitlines()
        csv_reader = csv.DictReader(decoded_file)

        # Loop through each row in the CSV
        for row in csv_reader:
            speech = row['Speech']  # Assuming 'Speech' column exists in the CSV

            # Analyze sentiment and polarity using TextBlob
            polarity = TextBlob(speech).sentiment.polarity
            print(polarity)
            sentiment = get_sentiment(speech)  # Assuming you have a function to determine sentiment

            # Store the data in the database
            analyzed_entry = SpeechModel(
                user_id=row['User'],  # Assuming 'User' column exists
                names=row['Name'],
                dob=row['DOB'],
                age=row['Age'],
                gender=row['Gender'],
                ethnicity=row['Ethnicity'],
                religion=row['Religion'],
                color=row['Color'],
                nationality=row['Nationality'],
                city=row['City'],
                speecher=row['Speecher'],
                location=row['Location'],
                speech=speech,
                status=sentiment,
                score=polarity
            )
            analyzed_entry.save()

            # Add the analyzed data to the list to display in the result page
            analyzed_data.append({
                "user": row['User'],
                "name": row['Name'],
                "dob": row['DOB'],
                "age": row['Age'],
                "gender": row['Gender'],
                "ethnicity": row['Ethnicity'],
                "religion": row['Religion'],
                "color": row['Color'],
                "nationality": row['Nationality'],
                "city": row['City'],
                "speecher": row['Speecher'],
                "location": row['Location'],
                "speech": speech,
                "sentiment": sentiment,
                "polarity": polarity
            })

        # Render the result.html page with the analyzed data
        return render(request, "result.html", {"data": analyzed_data})

    return render(request, "upload_dataset.html", {})


import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO
import base64
from django.shortcuts import render
from .models import SpeechModel


def generate_polarity_graphs():
    # Step 1: Fetch data from the SpeechModel table
    speeches = SpeechModel.objects.all()

    # Extract polarity scores and statuses
    polarity_scores = [speech.score for speech in speeches if speech.score is not None and not pd.isna(speech.score)]
    statuses = [speech.status for speech in speeches]

    # Convert data into a DataFrame
    df = pd.DataFrame({
        'polarity': polarity_scores,
        'status': statuses
    })

    # 1. Histogram of Polarity Scores (Distribution of Scores)
    plt.figure(figsize=(8, 5))
    sns.histplot(df["polarity"], bins=20, kde=True, color="skyblue")
    plt.title("Distribution of Polarity Scores")
    plt.xlabel("Polarity Score")
    plt.ylabel("Frequency")
    hist_chart = BytesIO()
    plt.savefig(hist_chart, format="png")
    hist_chart.seek(0)
    hist_chart_data = base64.b64encode(hist_chart.getvalue()).decode("utf-8")
    plt.close()

    # 2. Bar Chart for Polarity Status Categories (Positive, Neutral, Negative, Hate)
    status_counts = df['status'].value_counts()

    plt.figure(figsize=(8, 5))
    sns.barplot(x=status_counts.index, y=status_counts.values, palette='Blues_d')
    plt.title("Polarity Status Distribution")
    plt.xlabel("Polarity Status")
    plt.ylabel("Count")
    bar_chart = BytesIO()
    plt.savefig(bar_chart, format="png")
    bar_chart.seek(0)
    bar_chart_data = base64.b64encode(bar_chart.getvalue()).decode("utf-8")
    plt.close()

    # 3. Scatter Plot for Polarity Scores (showing polarity score for each speech)
    plt.figure(figsize=(8, 5))
    sns.scatterplot(x=range(len(df)), y=df['polarity'], color="orange")
    plt.title("Polarity Scores for Each Speech")
    plt.xlabel("Speech Index")
    plt.ylabel("Polarity Score")
    scatter_chart = BytesIO()
    plt.savefig(scatter_chart, format="png")
    scatter_chart.seek(0)
    scatter_chart_data = base64.b64encode(scatter_chart.getvalue()).decode("utf-8")
    plt.close()

    # 4. Scatter Plot with Status (Categorizing the scores by Status)
    category_map = {"Positive Speech": 1, "Neutral Speech": 2, "Negative Speech": 3, "Hate Speech": 4}
    df['category_numerical'] = df['status'].map(category_map)

    plt.figure(figsize=(8, 5))
    sns.scatterplot(x=range(len(df)), y=df['category_numerical'], hue=df['status'], palette='coolwarm', s=100)
    plt.title("Polarity Status vs. Polarity Scores")
    plt.xlabel("Speech Index")
    plt.ylabel("Polarity Status (Numerical)")
    status_scatter_chart = BytesIO()
    plt.savefig(status_scatter_chart, format="png")
    status_scatter_chart.seek(0)
    status_scatter_chart_data = base64.b64encode(status_scatter_chart.getvalue()).decode("utf-8")
    plt.close()

    return hist_chart_data, bar_chart_data, scatter_chart_data, status_scatter_chart_data


# View function to render the result page with the generated graphs
def display_polarity_graphs(request):
    # Step 1: Generate the polarity graphs
    hist_chart_data, bar_chart_data, scatter_chart_data, status_scatter_chart_data = generate_polarity_graphs()

    # Step 2: Render the result page with the graphs
    return render(request, "graph.html", {
        "hist_chart_data": hist_chart_data,
        "bar_chart_data": bar_chart_data,
        "scatter_chart_data": scatter_chart_data,
        "status_scatter_chart_data": status_scatter_chart_data
    })


def view_my_speech(request):
    email = request.session['email']
    obj = Registration.objects.get(email=email)
    obj1 = Speech.objects.filter(user_id=obj.id)
    scores = [speech.score for speech in obj1]
    labels = [speech.names for speech in obj1]  # Use names or any identifier for labels

    # Plotting the graph
    plt.figure(figsize=(10, 6))
    plt.plot(labels, scores, marker='o', linestyle='-', color='b')
    plt.title("Speech Scores Line Graph")
    plt.xlabel("Speech Name")
    plt.ylabel("Score (Polarity)")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()

    # Save plot to a string buffer
    buffer = io.BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    image_png = buffer.getvalue()
    buffer.close()

    # Convert plot to base64 to render in HTML
    graph = base64.b64encode(image_png).decode('utf-8')
    return render(request, 'view_my_speech.html', {"obj": obj1, "graph": graph})


def view_all_speech(request):
    email = request.session['email']
    obj = Registration.objects.get(email=email)
    obj1 = Speech.objects.all().exclude(user_id=obj.id)
    scores = [speech.score for speech in obj1]
    labels = [speech.names for speech in obj1]  # Use names or any identifier for labels

    # Plotting the graph
    plt.figure(figsize=(10, 6))
    plt.plot(labels, scores, marker='o', linestyle='-', color='b')
    plt.title("Speech Scores Line Graph")
    plt.xlabel("Speech Name")
    plt.ylabel("Score (Polarity)")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()

    # Save plot to a string buffer
    buffer = io.BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    image_png = buffer.getvalue()
    buffer.close()

    # Convert plot to base64 to render in HTML
    graph = base64.b64encode(image_png).decode('utf-8')
    return render(request, 'view_all_speech.html', {"obj": obj1, "graph": graph})


def admin_view_speech(request):
    obj1 = Speech.objects.all()
    scores = [speech.score for speech in obj1]
    labels = [speech.names for speech in obj1]  # Use names or any identifier for labels

    # Plotting the graph
    plt.figure(figsize=(10, 6))
    plt.plot(labels, scores, marker='o', linestyle='-', color='b')
    plt.title("Speech Scores Line Graph")
    plt.xlabel("Speech Name")
    plt.ylabel("Score (Polarity)")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()

    # Save plot to a string buffer
    buffer = io.BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    image_png = buffer.getvalue()
    buffer.close()

    # Convert plot to base64 to render in HTML
    graph = base64.b64encode(image_png).decode('utf-8')
    return render(request, 'admin_view_speech.html', {"obj": obj1, "graph": graph})


def _update(request):
    email = request.session["email"]
    return render(request, "_update.html", {})


def v_update(request):
    if request.method == "POST":
        id = request.POST["id"]
        user = Vendor.objects.get(id=id)
        form = VendorForm(request.POST, request.FILES, instance=user)
        print(form.errors)
        if form.is_valid():
            form.save()
            return redirect('/_my_profile')
    return redirect('/_my_profile')


def _home(request):
    return render(request, '_home.html', {})


def dai_login(request):
    if request.session.__contains__("email"):
        return True
    else:
        return False


def _change_password(request):
    email = request.session["email"]
    if dai_login(request):
        if request.method == "POST":
            password = request.POST["password"]
            new_password = request.POST["new_password"]
            try:
                user = Vendor.objects.get(email=email, password=password)
                user.password = new_password
                user.save()
                return redirect("/_login")
            except Exception as e:
                print(e)
                return render(request, "_change_password.html",
                              {"msg": "Your Password Not Changed, Please Try Again", "email": email})
        return render(request, "_change_password.html", {"email": email})
    return render(request, "_change_password.html", {"email": email})


def _login(request):
    if request.method == "POST":
        email = request.POST["email"]
        password = request.POST["password"]
        print(email, "", password)
        login = Vendor.objects.filter(email=email, password=password)
        if login.exists():
            print("hi")
            if login[0].status == "Accepted":
                request.session['email'] = email
                client = Vendor.objects.get(email=email)
                return render(request, "_home.html", {"msg": "customer.fullname"})
            else:
                return render(request, "_login.html", {"msg": "Account Is On Hold!"})
        else:
            return render(request, "_login.html", {"msg": "Invalid Email Or Password"})
    return render(request, "_login.html", {})


def _registration(request):
    if request.method == "POST":
        form = VendorForm(request.POST, request.FILES)
        print(form.errors)
        email = request.POST["email"]
        if Vendor.objects.filter(email=email).exists():
            return render(request, "_registration.html", {"msg": "This Email Is Already Exists "})
        else:
            if form.is_valid():
                form.save()
                return render(request, "_login.html", {"msg": "Your Details Are Registered Successfully"})
            return render(request, "_registration.html", {"msg": "Registration Failed"})
    return render(request, "_registration.html", {})


def accept_(request, id):
    accept = Vendor.objects.get(id=id)
    accept.status = 'Accepted'
    accept.save()
    return redirect('/admin_view_')


def reject_(request, id):
    reject = Vendor.objects.get(id=id)
    reject.status = 'Rejected'
    reject.save()
    return redirect('/admin_view_')


def admin_view_(request):
    ven = Vendor.objects.all()
    return render(request, 'admin_view_.html', {"ven": ven})


def admin_add_notifications(request):
    if request.method == "POST":
        form = NotificationsForm(request.POST)
        print(form.errors)
        if form.is_valid():
            form.save()
            return redirect('/admin_view_notification')
        return render(request, "admin_add_notifications.html", {"msg": "Notification Not Added"})
    return render(request, "admin_add_notifications.html", {})


def admin_logout(request):
    request.session.flush()
    return redirect('/admin_loginn')


def customer_view_notifications(request):
    note = Notifications.objects.all()
    return render(request, 'user_view_notifications.html', {"note": note})


def admin_view_notification(request):
    note = Notifications.objects.all()
    return render(request, 'admin_view_notifications.html', {"note": note})


def _view_notifications(request):
    dnote = Notifications.objects.all()
    return render(request, '_view_notifications.html', {"dnote": dnote})


def admin_delete_not(request, id):
    cus = Notifications.objects.get(id=id)
    cus.delete()
    return redirect('/admin_view_notification')


def admin_home(request):
    return render(request, 'admin_home.html', {})


def admin_view_contact(request):
    cont = Contact.objects.all()
    return render(request, 'admin_view_contact.html', {"cont": cont})


def admin_view_customer(request):
    con = Registration.objects.all()
    return render(request, 'admin_view_user.html', {"con": con})


def u_update(request):
    if request.method == "POST":
        id = request.POST["id"]
        user = Registration.objects.get(id=id)
        form = RegistrationForm(request.POST, request.FILES, instance=user)
        print(form.errors)
        if form.is_valid():
            form.save()
            return redirect('/my_profile')
    return redirect('/my_profile')


def my_profile(request):
    email = request.session["email"]
    pro = Registration.objects.get(email=email)
    return render(request, "my_profile.html", {"x": pro})


def c_update(request):
    email = request.session["email"]
    upd = Registration.objects.get(email=email)
    return render(request, "user_update.html", {"x": upd})


def about(request):
    return render(request, 'about.html', {})


def customer_home(request):
    return render(request, 'user_home.html', {})


def logout(request):
    request.session.flush()
    return redirect('/index')


def is_login(request):
    if request.session.__contains__("email"):
        return True
    else:
        return False


def delete_speech(request, id):
    obj = Speech.objects.get(id=id)
    obj.delete()
    return redirect('/view_my_speech')


def customer_change_password(request):
    email = request.session["email"]
    if is_login(request):
        if request.method == "POST":
            password = request.POST["password"]
            new_password = request.POST["new_password"]
            if password==new_password:
                return render(request, "user_change_password.html",
                              {"msg": "Current Password And New Password Should Not Be Same.", "email": email})
            else:
                try:
                    user = Registration.objects.get(email=email, password=password)
                    user.password = new_password
                    user.save()
                    return redirect('/customer_login')
                except Exception as e:
                    print(e)
                    return render(request, "user_change_password.html",
                                  {"msg": "Your Password Not Change, Please Try Again ", "email": email})
        return render(request, "user_change_password.html", {"email": email})
    return render(request, "user_change_password.html", {"email": email})


def admin_is_login(request):
    if request.session.__contains__("email"):
        return True
    else:
        return False


def contact(request):
    if request.method == "POST":
        form = ContactForm(request.POST)
        print(form.errors)
        if form.is_valid():
            form.save()
            return render(request, "contact.html", {'msg': 'Your Contact Is Register Successfully'})
        return render(request, "contact.html", {})
    return render(request, "contact.html", {})


def registration(request):
    if request.method == "POST":
        form = RegistrationForm(request.POST, request.FILES)
        print(form.errors)
        email = request.POST["email"]
        if Registration.objects.filter(email=email).exists() or Admin.objects.filter(email=email).exists():
            return render(request, "registration.html", {"msg": "This Email Is Already Exists"})
        else:
            if form.is_valid():
                form.save()
                return render(request, "user_login.html", {"msg": "Your Details Is Register Successfully"})
    return render(request, "registration.html", {})


# def customer_login(request):
#     if request.method == "POST":
#         email = request.POST["email"]
#         password = request.POST["password"]
#         log = Registration.objects.filter(email=email, password=password)
#         if log.exists():
#             request.session["email"] = email
#             return render(request, "user_home.html", {"msg": "Successfully login"})
#         else:
#             return render(request, "user_login.html", {"msg": "Incorrect Username or Password"})
#
#     return render(request, "user_login.html", {})

def customer_login(request):
    if request.method == "POST":
        email = request.POST["email"]
        password = request.POST["password"]
        print(email, "", password)
        login = Registration.objects.filter(email=email, password=password)
        if login.exists():
            print("hi")
            if login[0].status == "Accepted":
                request.session['email'] = email
                client = Registration.objects.get(email=email)
                return render(request, "user_home.html", {"msg": "Login Successfully"})
            else:
                return render(request, "user_login.html", {"msg": "Your Account Is On Hold !"})
        else:
            return render(request, "user_login.html", {"msg": "Invalid Data"})
    return render(request, "user_login.html", {})


def accept_customer(request, id):
    accept = Registration.objects.get(id=id)
    accept.status = 'Accepted'
    accept.save()
    return redirect('/admin_view_customer')


def reject_customer(request, id):
    reject = Registration.objects.get(id=id)
    reject.status = 'Rejected'
    reject.save()
    return redirect('/admin_view_customer')


def admin_loginn(request):
    if request.method == "POST":
        email = request.POST["email"]
        password = request.POST["password"]
        log = Admin.objects.filter(email=email, password=password)
        try:
            if log.exists():
                request.session["email"] = email
                return render(request, "admin_home.html", {"msg": "Successfully Login"})
            return render(request, "admin_loginn.html", {"msg": "Incorrect Email Or Password"})
        except Exception as e:
            print(e)
            return render(request, "admin_home.html", {"msg": ""})
    return render(request, "admin_loginn.html", {"msg": ""})


# Create your views here.
def admin_change_password(request):
    email = request.session["email"]
    if admin_is_login(request):
        print("hello")
        if request.method == 'POST':
            print("hello1")
            password = request.POST['password']
            new_password = request.POST['new_password']
            if password==new_password:
                return render(request, "admin_change_password.html",
                              {"msg": "Current Password And New Password Are Same.","email": email})
            else:

                try:
                    user = Admin.objects.get(email=email, password=password)
                    user.password = new_password
                    user.save()
                    return redirect('/admin_loginn')
                except Exception as e:
                    print(e)
                    return render(request, "admin_change_password.html",
                                  {"msg": "Your Password Not Changed, Please Try Again", "email": email})
        return render(request, "admin_change_password.html", {"email": email})
    return render(request, "admin_change_password.html", {"email": email})


def admin_view_(request):
    pro = _add_.objects.all()
    return render(request, "admin_view_.html", {"x": pro})


def customer_view_(request, email):
    cpro = _add_.objects.filter(email=email)
    return render(request, "customer_view_.html", {"x": cpro})


def _view_(request):
    email = request.session["email"]
    dpro = _add_.objects.filter(email=email)
    return render(request, "_view_.html", {"x": dpro})

#service provider
def serviceproviderlogin(request):
    if request.method == "POST":
        admin = request.POST.get('username')
        password = request.POST.get('password')
        if admin == "admin" and password == "admin":
            return redirect('View_PersonSpeechData_Details')

    return render(request, 'SProvider/serviceproviderlogin.html')


def viewtreandingquestions(request, chart_type):
    dd = {}
    pos, neu, neg = 0, 0, 0
    poss = None
    topic = hate_speech_review_model.objects.values('ratings').annotate(dcount=Count('ratings'))._by('-dcount')
    for t in topic:
        topics = t['ratings']
        pos_count = hate_speech_review_model.objects.filter(topics=topics).values('names').annotate(
            topiccount=Count('ratings'))
        poss = pos_count
        for pp in pos_count:
            senti = pp['names']
            if senti == 'positive':
                pos = pp['topiccount']
            elif senti == 'negative':
                neg = pp['topiccount']
            elif senti == 'nutral':
                neu = pp['topiccount']
        dd[topics] = [pos, neg, neu]
    return render(request, 'SProvider/viewtreandingquestions.html',
                  {'object': topic, 'dd': dd, 'chart_type': chart_type})


def Search_Person(request):  # Search
    if request.method == "POST":
        kword = request.POST.get('keyword')
        obj = hate_speech_review_model.objects.all().filter(Q(names__contains=kword))
        return render(request, 'SProvider/Search_Person.html', {'list_objects': obj})
    return render(request, 'SProvider/Search_Person.html')


def View_HateSpeech_Speech(request):  # Search
    f1 = 'detest'
    f2 = 'hate'
    f3 = 'disgusting'
    f4 = 'disgust'
    f5 = 'dislike'
    f6 = 'hated'
    f11 = 'outrage'
    f22 = 'killed'

    obj1 = hate_speech_review_model.objects.all()
    obj = hate_speech_review_model.objects.all().filter(
        Q(speech__contains=f1) | Q(speech__contains=f2) | Q(speech__contains=f3) | Q(speech__contains=f4) |
        Q(speech__contains=f5) | Q(speech__contains=f6) | Q(speech__contains=f11) | Q(speech__contains=f22))

    count = obj.count()
    count1 = obj1.count()
    hatespeech = (count / count1) * 100

    return render(request, 'SProvider/View_HateSpeech_Speech.html', {'list_objects': obj, 'count': hatespeech})


def View_Positive_Speech(request):  # Positive

    f1 = 'good'
    f2 = 'beautiful'
    f3 = 'fantastic'
    f4 = 'extraordinary'
    f5 = 'best'
    f6 = 'healthy'
    f7 = 'happy'
    f8 = 'marvel'
    f9 = 'worth'
    f10 = 'value'
    f11 = 'amazing'
    obj = hate_speech_review_model.objects.all().filter(
        Q(speech__contains=f1) | Q(speech__contains=f2) | Q(speech__contains=f3) | Q(speech__contains=f4) | Q(
            speech__contains=f5) | Q(speech__contains=f6) | Q(speech__contains=f7) | Q(speech__contains=f8) | Q(
            speech__contains=f9) | Q(speech__contains=f10) | Q(speech__contains=f11))
    obj1 = hate_speech_review_model.objects.all()
    count = obj.count()
    count1 = obj1.count()
    positive = (count / count1) * 100
    return render(request, 'SProvider/View_Positive_Speech.html', {'list_objects': obj, 'count': positive})


def View_Negative_Speech(request):
    f1 = 'bad'
    f2 = 'worst'
    f3 = 'heavy'
    f4 = 'ridicules'
    f5 = 'sad'
    f6 = 'damn'
    f7 = 'shameful'
    f8 = 'failed'
    obj = hate_speech_review_model.objects.all().filter(
        Q(speech__contains=f1) | Q(speech__contains=f2) | Q(speech__contains=f3) | Q(speech__contains=f4) | Q(
            speech__contains=f5) | Q(speech__contains=f6) | Q(speech__contains=f7) | Q(speech__contains=f8))
    obj1 = hate_speech_review_model.objects.all()
    count = obj.count()
    count1 = obj1.count()
    negative = (count / count1) * 100
    return render(request, 'SProvider/View_Negative_Speech.html', {'list_objects': obj, 'count': negative})


def View_Remote_Users(request):
    obj = ClientRegister_Model.objects.all()
    return render(request, 'SProvider/View_Remote_Users.html', {'objects': obj})


def ViewTrendings(request):
    topic = hate_speech_review_model.objects.values('topics').annotate(dcount=Count('topics'))._by('-dcount')
    return render(request, 'SProvider/ViewTrendings.html', {'objects': topic})


def negativechart(request, chart_type):
    dd = {}
    pos, neu, neg = 0, 0, 0
    poss = None
    topic = hate_speech_review_model.objects.values('ratings').annotate(dcount=Count('ratings'))._by('-dcount')
    for t in topic:
        topics = t['ratings']
        pos_count = hate_speech_review_model.objects.filter(topics=topics).values('names').annotate(
            topiccount=Count('ratings'))
        poss = pos_count
        for pp in pos_count:
            senti = pp['names']
            if senti == 'positive':
                pos = pp['topiccount']
            elif senti == 'negative':
                neg = pp['topiccount']
            elif senti == 'nutral':
                neu = pp['topiccount']
        dd[topics] = [pos, neg, neu]
    return render(request, 'SProvider/negativechart.html', {'object': topic, 'dd': dd, 'chart_type': chart_type})


def charts(request, chart_type):
    chart1 = hate_speech_review_model.objects.values('names').annotate(dcount=Avg('score'))
    return render(request, "SProvider/charts.html", {'form': chart1, 'chart_type': chart_type})


def View_PersonSpeechDataSets_Details(request):
    obj = hate_speech_review_model.objects.all()
    return render(request, 'SProvider/View_PersonSpeechDataSets_Details.html', {'list_objects': obj})


def likeschart(request, like_chart):
    charts = hate_speech_review_model.objects.values('names').annotate(dcount=Avg('score'))
    return render(request, "SProvider/likeschart.html", {'form': charts, 'like_chart': like_chart})


def View_PersonSpeechData_Details(request):
    obj = Add_Person_Data_Model.objects.all()
    return render(request, 'SProvider/View_PersonSpeechData_Details.html', {'list_objects': obj})


def add_notifications(request):
    if request.method == "POST":
        form = Add_notifciationsforms(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/view_notifications')
        return render(request, 'SProvider/add_notifications.html', {"msg": "Failed To Add"})
    return render(request, 'SProvider/add_notifications.html', {})


def view_notifications(request):
    notifications = Add_notications.objects.all()
    return render(request, 'SProvider/view_notifications.html', {'notifications': notifications})


def delete_notifications(request, id):
    notifications = Add_notications.objects.get(id=id)
    notifications.delete()
    return redirect('view_notifications')


def logout(request):
    request.session.flush()
    return redirect('/serviceproviderlogin')


#user views

from django.db.models import Count
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
import datetime
import openpyxl
from textblob import TextBlob


def login(request):
    if request.method == "POST" and 'submit1' in request.POST:

        username = request.POST.get('username')
        password = request.POST.get('password')
        try:

            enter = ClientRegister_Model.objects.get(username=username, password=password)
            request.session["userid"] = enter.id
            return redirect('Add_DataSet_Details')
        except Exception as e:
            print(e)
            return render(request, 'RUser/login.html')
    return render(request, 'RUser/login.html')


def Add_DataSet_Details(request):
    if "GET" == request.method:
        return render(request, 'RUser/Add_DataSet_Details.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)

            hate_speech_review_model.objects.all().delete()

    for r in range(1, active_sheet.max_row + 1):
        hate_speech_review_model.objects.create(
            names=active_sheet.cell(r, 1).value,
            age=active_sheet.cell(r, 2).value,
            dob=active_sheet.cell(r, 3).value,
            gender=active_sheet.cell(r, 4).value,
            ethnicity=active_sheet.cell(r, 5).value,
            religion=active_sheet.cell(r, 6).value,
            color=active_sheet.cell(r, 7).value,
            nationality=active_sheet.cell(r, 8).value,
            city=active_sheet.cell(r, 9).value,
            speecher=active_sheet.cell(r, 10).value,
            speech=active_sheet.cell(r, 11).value,
            sdate=active_sheet.cell(r, 12).value,
            location=active_sheet.cell(r, 13).value,
            score=active_sheet.cell(r, 14).value

        )

    return render(request, 'RUser/Add_DataSet_Details.html', {"excel_data": excel_data})


def Register1(request):
    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        phoneno = request.POST.get('phoneno')
        country = request.POST.get('country')
        state = request.POST.get('state')
        city = request.POST.get('city')
        ClientRegister_Model.objects.create(username=username, email=email, password=password, phoneno=phoneno,
                                            country=country, state=state, city=city)

        return render(request, 'RUser/Register1.html')
    else:

        return render(request, 'RUser/Register1.html')


def ViewYourProfile(request):
    userid = request.session['userid']
    obj = ClientRegister_Model.objects.get(id=userid)
    return render(request, 'RUser/ViewYourProfile.html', {'object': obj})


def Search_PersonDataSets(request):
    if request.method == "POST":
        kword = request.POST.get('keyword')
        obj = hate_speech_review_model.objects.all().filter(Q(names__contains=kword))
        return render(request, 'RUser/Search_PersonDataSets.html', {'list_objects': obj})
    return render(request, 'RUser/Search_PersonDataSets.html')


def ratings(request, pk):
    vott1, vott, neg = 0, 0, 0
    objs = hate_speech_review_model.objects.get(id=pk)
    unid = objs.id
    vot_count = hate_speech_review_model.objects.all().filter(id=unid)
    for t in vot_count:
        vott = t.ratings
        vott1 = vott + 1
        obj = get_object_or_404(hate_speech_review_model, id=unid)
        obj.ratings = vott1
        obj.save(update_fields=["ratings"])
        return redirect('Add_DataSet_Details')

    return render(request, 'RUser/ratings.html', {'objs': vott1})


def person(request):
    userid = request.session["userid"]
    data = Add_Person_Data_Model.objects.filter(userid=userid)
    if request.method == "POST":
        form = Personforms(request.POST)
        if form.is_valid():
            names = form.cleaned_data["names"]
            dob = form.cleaned_data["dob"]
            age = form.cleaned_data["age"]
            gender = form.cleaned_data["gender"]
            ethnicity = form.cleaned_data["ethnicity"]
            religion = form.cleaned_data["religion"]
            color = form.cleaned_data["color"]
            nationality = form.cleaned_data["nationality"]
            city = form.cleaned_data["city"]
            speecher = form.cleaned_data["speecher"]
            location = form.cleaned_data["location"]
            score = form.cleaned_data["score"]
            speech = form.cleaned_data["speech"]
            res = get_sentiment(speech)
            print(res)
            obj = Add_Person_Data_Model()
            obj.userid = userid
            obj.names = names
            obj.gender = gender
            obj.ethnicity = ethnicity
            obj.religion = religion
            obj.color = color
            obj.nationality = nationality
            obj.city = city
            obj.location = location
            obj.speecher = speecher
            obj.age = age
            obj.score = score
            obj.dob = dob
            obj.speech = speech
            obj.status = res
            obj.save()
        return render(request, "RUser/person.html", {"msg": "Data Posted", "data": data, "userid": userid})
    return render(request, "RUser/person.html", {"userid": userid, "data": data})


# def Search_Person_manual_data(request):
#     userid = request.session["userid"]
#     queryval = request.POST["data"]
#     data = Add_Person_Data_Model.objects.filter(names__contains=queryval | Add_Person_Data_Model.objects.filter(speecher__contains=queryval)
#     return render(request,"RUser/person.html",{"data":data})

def Search_Person_manual_data(request):
    userid = request.session["userid"]
    queryval = request.POST["data"]
    data = Add_Person_Data_Model.objects.filter(
        Q(names__icontains=queryval) | Q(speecher__icontains=queryval)
    )
    return render(request, "RUser/person.html", {"data": data})


def delete_manual_data(request, id):
    userid = request.session["userid"]
    data = Add_Person_Data_Model.objects.get(id=id)
    data.delete()
    return redirect('/person')


def get_sentiment(speech):
    if len(speech) == 0:
        return "Empty"
    res = TextBlob(speech)
    sentiment = res.sentiment.polarity
    if any(word in speech.lower() for word in [
        "hate", "dislike", "disgusting", "disgust", "detest", "hated", "outrage", "killed",
        "racist", "bigot", "intolerant", "supremacist", "oppressor", "ignorant", "violent",
        "hostile", "prejudice", "segregation", "offensive", "abhorrent", "demeaning",
        "evil", "vile", "exterminate", "vermin", "cockroach", "parasite", "inferior",
        "genocide", "lynch", "terrorist", "traitor", "inhuman", "barbaric", "slur",
        "threat", "murderer", "criminal", "corrupt", "worthless", "uncivilized",
        "backward", "savage", "devil", "demon", "unfit", "unworthy", "trash", "garbage",
        "scum", "vermin", "plague", "enemy", "traitor", "invader", "destroy", "obliterate",
        "brute", "coward", "pig", "dog", "monster", "horrible", "abomination", "horrific",
        "menace", "despicable", "filthy", "unclean"
    ]):
        return "hate speech"
    elif sentiment > 0 and sentiment <= 1:
        return "positive speech"
    elif sentiment >= -1 and sentiment < 0:
        return "negative speech"
    else:
        return "neutral speech"


def edit_persondata(request, id):
    data = Add_Person_Data_Model.objects.get(id=id)
    return render(request, "RUser/edit_persondata.html", {"data": data})


def update_person(request):
    if request.method == "POST":
        id = request.POST["id"]
        userid = request.POST["userid"]
        data_instance = Add_Person_Data_Model.objects.get(id=id, userid=userid)
        form = Personforms(request.POST, instance=data_instance)
        if form.is_valid():
            speech = form.cleaned_data['speech']
            sentiment_status = get_sentiment(speech)
            data_instance.status = sentiment_status
            form.save()
            return redirect("/person")
    return redirect("/person")


def view_notification(request):
    notifications = Add_notications.objects.all()
    return render(request, 'RUser/view_notification.html', {'notifications': notifications})


def update_profile(request):
    id = request.session['userid']
    user = ClientRegister_Model.objects.get(id=id)
    if request.method == 'POST':
        form = ClientRegister_Form(request.POST)
        if form.is_valid():
            obj = ClientRegister_Model()
            obj.id = id
            obj.username = form.cleaned_data['username']
            obj.email = user.email
            obj.password = user.password
            obj.phoneno = form.cleaned_data['phoneno']
            obj.country = form.cleaned_data['country']
            obj.state = form.cleaned_data['state']
            obj.city = form.cleaned_data['city']
            obj.save()
            return redirect('/ViewYourProfile')
        return render(request, 'RUser/update_profile.html', {'x': user})
    return render(request, 'RUser/update_profile.html', {'x': user})
